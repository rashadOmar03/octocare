import csv
import io
from pathlib import Path
from typing import Iterable

from fastapi.responses import StreamingResponse


def _escape_xml(text: str) -> str:
    return (
        str(text or "")
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
    )


def stream_csv(rows: Iterable[list], filename: str) -> StreamingResponse:
    buf = io.StringIO()
    writer = csv.writer(buf)
    for row in rows:
        writer.writerow(row)
    data = buf.getvalue().encode("utf-8-sig")
    return StreamingResponse(
        io.BytesIO(data),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def stream_xlsx(sheets: dict[str, list[list]], filename: str) -> StreamingResponse:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError(
            "XLSX export requires openpyxl. Run: python -m pip install openpyxl==3.1.5"
        ) from exc

    wb = Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill("solid", fgColor="2C3E50")
    header_font = Font(bold=True, color="FFFFFF", size=10)

    for sheet_name, rows in sheets.items():
        ws = wb.create_sheet(title=sheet_name[:31])
        if not rows:
            continue
        for r_idx, row in enumerate(rows, start=1):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                else:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
        for col_idx in range(1, len(rows[0]) + 1):
            letter = get_column_letter(col_idx)
            max_len = 12
            for row in rows[:200]:
                if col_idx - 1 < len(row):
                    max_len = max(max_len, min(len(str(row[col_idx - 1] or "")), 48))
            ws.column_dimensions[letter].width = max_len + 2

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _write_proof_sheet(ws, entries: list[dict], header_fill, header_font) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    ws.append(["Reference", "Proof File", "Stored Path"])
    for col_idx in range(1, 4):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.column_dimensions["A"].width = 52
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 28

    image_row_height = 180
    for entry in entries:
        row_idx = ws.max_row + 1
        ws.cell(row=row_idx, column=1, value=entry.get("label", ""))
        ws.cell(row=row_idx, column=3, value=entry.get("url", ""))
        ws.row_dimensions[row_idx].height = image_row_height
        for col_idx in (1, 3):
            ws.cell(row=row_idx, column=col_idx).alignment = Alignment(vertical="top", wrap_text=True)

        path = entry.get("path")
        path_str = str(path) if path is not None else ""
        if path is not None and Path(path_str).is_file() and path_str.lower().endswith(
            (".jpg", ".jpeg", ".png", ".gif", ".webp", ".bmp")
        ):
            try:
                from openpyxl.drawing.image import Image as XLImage

                img = XLImage(path_str)
                img.width = min(img.width, 320)
                img.height = min(img.height, 220)
                ws.add_image(img, f"B{row_idx}")
                ws.cell(row=row_idx, column=2, value="Embedded below")
            except Exception:
                ws.cell(row=row_idx, column=2, value="Image embed failed")
        else:
            ws.cell(row=row_idx, column=2, value="No image file on server")


def stream_xlsx_with_proofs(
    sheets: dict[str, list[list]],
    payment_proofs: list[dict],
    refund_proofs: list[dict],
    filename: str,
) -> StreamingResponse:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError(
            "XLSX export requires openpyxl. Run: python -m pip install openpyxl==3.1.5"
        ) from exc

    wb = Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill("solid", fgColor="2C3E50")
    header_font = Font(bold=True, color="FFFFFF", size=10)

    for sheet_name, rows in sheets.items():
        ws = wb.create_sheet(title=sheet_name[:31])
        if not rows:
            continue
        for r_idx, row in enumerate(rows, start=1):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                else:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
        for col_idx in range(1, len(rows[0]) + 1):
            letter = get_column_letter(col_idx)
            max_len = 12
            for row in rows[:200]:
                if col_idx - 1 < len(row):
                    max_len = max(max_len, min(len(str(row[col_idx - 1] or "")), 48))
            ws.column_dimensions[letter].width = max_len + 2

    if payment_proofs:
        ws_pay = wb.create_sheet(title="Payment Proofs"[:31])
        _write_proof_sheet(ws_pay, payment_proofs, header_fill, header_font)

    if refund_proofs:
        ws_ref = wb.create_sheet(title="Refund Proofs"[:31])
        _write_proof_sheet(ws_ref, refund_proofs, header_fill, header_font)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
